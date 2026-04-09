import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ================= 配置路径 =================
data_dir = '/home/wugang/Data/Orthopedics_Dataset/Hospital_Dataset/Table_Dataset/Resort_Raw_Tables'
temp_dir = '../Temp_data' 
mapping_file = os.path.join(temp_dir, '30_Lab_Items_Mapping_For_Doctors_V6_Final.xlsx')
# 新增：碰撞报告的 Excel 导出路径
report_out_file = os.path.join(temp_dir, '31_Collision_Detection_Report.xlsx')

print("🚀 启动【全量检验指标动态碰撞探针】(含 Excel 自动导出)...\n")

# 1. 加载映射字典 (提取绿区和黄区的合并组)
df_map = pd.read_excel(mapping_file)
merged_groups = df_map[~df_map['模块分类'].str.contains('三、独立', na=False)].copy()

target_dict = {} 
for _, row in merged_groups.iterrows():
    std_name = row['推荐标准名称']
    details_str = row['包含的原始明细变体']
    raw_names = [re.sub(r' \(频次:\d+\)', '', item).strip() for item in str(details_str).split('|')]
    if len(raw_names) > 1:
        target_dict[std_name] = raw_names

print(f"📊 待检查的潜在合并组合数量: {len(target_dict)} 个")

# 2. 遍历读取原始检验大表
lab_files = ['Lab_Results_1.xlsx', 'Lab_Results_2.xlsx', 'Lab_Results_3.xlsx']
all_lab_data = []

print("📂 正在扫描原始检验大数据池...")
for file in lab_files:
    file_path = os.path.join(data_dir, file)
    if not os.path.exists(file_path): continue
    try:
        df = pd.read_excel(file_path, usecols=['患者编号', '报告时间', '分析明细'])
        all_lab_data.append(df)
    except Exception as e:
        print(f"读取 {file} 出错: {e}")

if not all_lab_data:
    print("未找到原始检验数据，请检查路径。")
    exit()

big_df = pd.concat(all_lab_data, ignore_index=True)
big_df['分析明细'] = big_df['分析明细'].astype(str).str.strip()

# 3. 开始执行共现碰撞检查
print("\n🔍 正在进行交叉火力扫描 (这可能需要十几秒)...\n")

collision_reports = []
safe_count = 0

for std_name, raw_items in target_dict.items():
    subset = big_df[big_df['分析明细'].isin(raw_items)].copy()
    if subset.empty: continue
    
    subset['Event_ID'] = subset['患者编号'].astype(str) + "_" + subset['报告时间'].astype(str)
    event_counts = subset.groupby('Event_ID')['分析明细'].nunique()
    
    collisions = event_counts[event_counts > 1]
    
    if len(collisions) > 0:
        collision_rate = len(collisions) / len(event_counts)
        collision_details = subset[subset['Event_ID'].isin(collisions.index)]
        involved_pairs = collision_details.groupby('Event_ID')['分析明细'].apply(lambda x: " + ".join(sorted(x.unique()))).value_counts()
        
        collision_reports.append({
            '推荐标准名称': std_name,
            '总出现次数(事件数)': len(event_counts),
            '碰撞次数': len(collisions),
            '碰撞率': f"{collision_rate*100:.1f}%",
            '具体的致命碰撞组合': involved_pairs.index[0] 
        })
    else:
        safe_count += 1

# 4. 输出终端诊断报告并保存为 Excel
print(f"✅ 扫描完毕！在 {len(target_dict)} 个合并组中，有 【 {safe_count} 】 个是绝对安全的！")

if collision_reports:
    report_df = pd.DataFrame(collision_reports).sort_values(by='碰撞次数', ascending=False)
    
    # 终端打印
    print("\n🚨 警告！抓到了以下几个伪装成合并项的“雷区”：")
    print("-" * 100)
    print(report_df.to_string(index=False))
    print("-" * 100)
    
    # =============== 核心新增：导出为带格式的 Excel ===============
    wb = Workbook()
    ws = wb.active
    ws.title = "同源碰撞警告报告"
    
    # 将 DataFrame 写入 Excel
    for r in dataframe_to_rows(report_df, index=False, header=True):
        ws.append(r)
        
    # 简单的格式美化（表头加粗，背景标红警告）
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="9C0006")
        cell.fill = danger_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50

    wb.save(report_out_file)
    print(f"\n💾 【重大进展】碰撞诊断报告已成功导出为 Excel！请查看: {report_out_file}")
    print("👉 请打开此表格，遇到里面出现的组合，必须在下一次长表转宽表时将其拆解或设定优先级保留策略！")
else:
    print("\n🎉 奇迹出现！你准备的所有合并组在全量数据下都没有发生任何碰撞！")
    # 即使没有碰撞，也生成一个包含"无碰撞"结论的空表以作科研记录
    pd.DataFrame({'诊断结果': ['全量数据扫描通过，未发现任何同源合并冲突']}).to_excel(report_out_file, index=False)
    print(f"💾 安全审查通过记录已保存至: {report_out_file}")