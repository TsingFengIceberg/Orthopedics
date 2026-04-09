import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ================= 配置路径 =================
temp_dir = '../Temp_data' 
input_file = os.path.join(temp_dir, '26_Lab_Items_Full_Dictionary.csv')
output_file = os.path.join(temp_dir, '29_Lab_Items_Mapping_For_Doctors.xlsx')

print("🚀 正在构建面向医生的【可视化审查版】映射字典...\n")

# ================= 1. 规则定义区 =================
# 【绿色阵营】：100% 确认安全的合并 (包含新增的细胞计数、尿素肌酐等)
SAFE_MERGE_DICT = {
    '乙型肝炎病毒表面抗原(CLIA)': '乙型肝炎病毒表面抗原', '乙型肝炎病毒表面抗原(ELISA)': '乙型肝炎病毒表面抗原',
    '丙型肝炎病毒抗体(CLIA)': '丙型肝炎病毒抗体', '丙型肝炎病毒抗体(ELISA)': '丙型肝炎病毒抗体',
    '梅毒特异性抗体(TP-ELISA)': '梅毒特异性抗体', '梅毒特异性抗体(CLIA)': '梅毒特异性抗体', '梅毒特异性抗体(TPPA)': '梅毒特异性抗体',
    '人类免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(ELISA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体',
    '乙型肝炎病毒e抗体(ELISA)': '乙型肝炎病毒e抗体', '乙型肝炎病毒e抗体(CLIA)': '乙型肝炎病毒e抗体',
    '乙型肝炎病毒e抗原(ELISA)': '乙型肝炎病毒e抗原', '乙型肝炎病毒e抗原(CLIA)': '乙型肝炎病毒e抗原',
    '乙型肝炎病毒核心抗体(ELISA)': '乙型肝炎病毒核心抗体', '乙型肝炎病毒核心抗体(CLIA)': '乙型肝炎病毒核心抗体',
    '乙型肝炎病毒表面抗体(ELISA)': '乙型肝炎病毒表面抗体', '乙型肝炎病毒表面抗体(CLIA)': '乙型肝炎病毒表面抗体',
    '乙型肝炎病毒核心抗体IgM(ELISA)': '乙型肝炎病毒核心抗体IgM', '乙型肝炎病毒核心抗体IgM(CLIA)': '乙型肝炎病毒核心抗体IgM',
    '总前列腺特异性抗原(tPSA)': '总前列腺特异性抗原', '总前列腺特异性抗原（tPSA）': '总前列腺特异性抗原',
    '游离前列腺特异性抗原(fPSA)': '游离前列腺特异性抗原', '游离前列腺特异性抗原（fPSA）': '游离前列腺特异性抗原',
    '胃蛋白酶原Ⅰ(PGⅠ)': '胃蛋白酶原Ⅰ', '胃蛋白酶原Ⅱ(PGⅡ)': '胃蛋白酶原Ⅱ',
    '总Ⅰ型(前)胶原氨基端延长肽': '总Ⅰ型胶原氨基端延长肽',
    '血沉（毛细管法）': '血沉', '血沉（光度计法）': '血沉',
    '国际标准化比率': '国际标准化比值', 'UN:CREA': 'BUN:CREA',
    'a-L-岩藻糖甘酶': 'a-L-岩藻糖苷酶', '病理性管型': '病理管型', '酵母样菌': '酵母菌',
    '胃泌素释放前肽前体': '胃泌素释放肽前体', '谷氨酰转酞酶': '谷氨酰转肽酶', '谷氨酰基转肽酶': '谷氨酰转肽酶',
    '亮氨酸氨基转肽酶': '亮氨酸氨基肽酶', '腺苷酸脱氨酶': '腺苷脱氨酶', '超敏C-反应蛋白': '超敏C反应蛋白',
    '糖类抗原153': '糖类抗原15-3', '糖类抗原199': '糖类抗原19-9',
    # 新增的 100% 安全合并
    '白细胞数': '白细胞', '白细胞计数': '白细胞',
    '红细胞数': '红细胞', '红细胞计数': '红细胞',
    '血小板数': '血小板', '血小板计数': '血小板',
    '尿素氮': '尿素', '尿素计算': '尿素',
    '肌酐计算': '肌酐'
}

# 【黄色阵营】：疑似可合并，但需医生定夺 (因公式、敏感度不同)
SUSPECT_MERGE_DICT = {
    '肾小球滤过率(MDRD)': '【存疑待确认】肾小球滤过率', '肾小球滤过率(eGFR-EPI)': '【存疑待确认】肾小球滤过率', '肾小球滤过率（女）': '【存疑待确认】肾小球滤过率', '肾小球滤过率(男)': '【存疑待确认】肾小球滤过率',
    '隐血试验（胶体金法）': '【存疑待确认】隐血试验', '隐血试验': '【存疑待确认】隐血试验', '隐血试验（匹拉米洞法）': '【存疑待确认】隐血试验', '隐血试验（化学法）': '【存疑待确认】隐血试验',
    '肌酸激酶同工酶（MB）': '【存疑待确认】肌酸激酶同工酶', '肌酸激酶同工酶（MB质量法）': '【存疑待确认】肌酸激酶同工酶', '肌酸激酶同工酶': '【存疑待确认】肌酸激酶同工酶',
    'ROMA值(绝经前)': '【存疑待确认】ROMA值', 'ROMA值(绝经后)': '【存疑待确认】ROMA值'
}

# ================= 2. 数据处理与映射 =================
df = pd.read_csv(input_file)

def extract_analyte(full_name):
    parts = str(full_name).split(' - ')
    if len(parts) >= 2: return ' - '.join(parts[1:]).strip()
    return str(full_name).strip()

df['原始明细'] = df['完整检验项 (项目 - 明细)'].apply(extract_analyte)
df['原始明细'] = df['原始明细'].str.strip(' *-_')

# 将同名（跨套餐）的直接累加频次
raw_counts = df.groupby('原始明细')['出现总频次'].sum().reset_index()

# 映射分类逻辑
def classify_and_map(name):
    if name in SAFE_MERGE_DICT: return SAFE_MERGE_DICT[name], '1_Safe'
    if name in SUSPECT_MERGE_DICT: return SUSPECT_MERGE_DICT[name], '2_Suspect'
    # 对于不在上面两个字典里，但是有 (CV)/(SD)/(镜检) 等明确防误杀标记的，加上分类标签
    if '(CV)' in name or '(SD)' in name or '镜检' in name or '胸腹水' in name or '脑脊液' in name:
        return name, '3_DoNotMerge'
    # 其他全部单项
    return name, '3_DoNotMerge'

raw_counts[['推荐标准名称', 'Category']] = raw_counts['原始明细'].apply(lambda x: pd.Series(classify_and_map(x)))

# 进行折叠统计
result_data = []
for (std_name, cat), group in raw_counts.groupby(['推荐标准名称', 'Category']):
    total_freq = group['出现总频次'].sum()
    details = "  |  ".join([f"{r['原始明细']} (频次:{r['出现总频次']})" for _, r in group.sort_values(by='出现总频次', ascending=False).iterrows()])
    result_data.append({
        'Category': cat, '推荐标准名称': std_name, '总合并频次': total_freq, 
        '合并项数量': len(group), '包含的原始明细变体': details
    })

final_df = pd.DataFrame(result_data)

# ================= 3. 写入 Excel 并上色 =================
wb = Workbook()
ws = wb.active
ws.title = "检验指标审查"

# 设置表头
headers = ['模块分类', '推荐标准名称', '总合并频次', '合并项数量', '包含的原始明细变体']
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 颜色配置（浅色柔和系）
color_safe = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # 浅绿
color_suspect = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄
color_danger = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # 浅灰（保留原样）

current_cat = ''
for cat in ['1_Safe', '2_Suspect', '3_DoNotMerge']:
    sub_df = final_df[final_df['Category'] == cat].sort_values(by=['合并项数量', '总合并频次'], ascending=[False, False])
    
    if len(sub_df) == 0: continue
    
    # 根据类别设定显示文本和颜色
    if cat == '1_Safe':
        cat_name = "【第一部分：100%安全合并 (可直接通过)】"
        fill_color = color_safe
    elif cat == '2_Suspect':
        cat_name = "【第二部分：存疑待定夺 (需医生评估临床意义)】"
        fill_color = color_suspect
    else:
        cat_name = "【第三部分：独立指标 / 绝对不能合并的 (保持现状)】"
        fill_color = color_danger
        
    # 添加一个空行做隔离
    ws.append(['', '', '', '', '']) 
    
    for _, row in sub_df.iterrows():
        ws.append([cat_name, row['推荐标准名称'], row['总合并频次'], row['合并项数量'], row['包含的原始明细变体']])
        # 给这一行上色
        for cell in ws[ws.max_row]:
            cell.fill = fill_color
            cell.alignment = Alignment(vertical='center')

# 调整列宽使其美观
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 80

wb.save(output_file)
print(f"🎉 审查版 Excel 生成完毕！已保存至: {output_file}")
print("👉 打开文件后，你会看到极其舒适的三色分区：绿色通过、黄色审阅、灰色保留。交给医生看这版准没错！")