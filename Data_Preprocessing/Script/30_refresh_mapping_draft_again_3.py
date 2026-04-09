import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ================= 配置路径 =================
temp_dir = '../Temp_data' 
input_file = os.path.join(temp_dir, '26_Lab_Items_Full_Dictionary.csv')
output_file = os.path.join(temp_dir, '30_Lab_Items_Mapping_For_Doctors_V6_Final.xlsx')

print("🚀 正在执行【跨区交叉核验】构建 V6 终极收官版字典...\n")

# ================= 1. 规则定义区 =================
SAFE_MERGE_DICT = {
    # 传染病与抗原抗体
    '乙型肝炎病毒表面抗原(CLIA)': '乙型肝炎病毒表面抗原', '乙型肝炎病毒表面抗原(ELISA)': '乙型肝炎病毒表面抗原',
    '丙肝抗体(CLIA)': '丙型肝炎病毒抗体', '丙型肝炎病毒抗体(CLIA)': '丙型肝炎病毒抗体', '丙型肝炎病毒抗体(ELISA)': '丙型肝炎病毒抗体',
    '梅毒特异性抗体(TP-ELISA)': '梅毒特异性抗体', '梅毒特异性抗体(CLIA)': '梅毒特异性抗体', '梅毒特异性抗体(TPPA)': '梅毒特异性抗体',
    '人类免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(ELISA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体',
    '艾滋病抗原抗体筛查(ELISA)': '人类免疫缺陷病毒抗原抗体', 'HIV P24抗原/抗体(CLIA)': '人类免疫缺陷病毒抗原抗体',
    '乙型肝炎病毒e抗体(ELISA)': '乙型肝炎病毒e抗体', '乙型肝炎病毒e抗体(CLIA)': '乙型肝炎病毒e抗体',
    '乙型肝炎病毒e抗原(ELISA)': '乙型肝炎病毒e抗原', '乙型肝炎病毒e抗原(CLIA)': '乙型肝炎病毒e抗原',
    '乙型肝炎病毒核心抗体(ELISA)': '乙型肝炎病毒核心抗体', '乙型肝炎病毒核心抗体(CLIA)': '乙型肝炎病毒核心抗体',
    '乙型肝炎病毒表面抗体(ELISA)': '乙型肝炎病毒表面抗体', '乙型肝炎病毒表面抗体(CLIA)': '乙型肝炎病毒表面抗体',
    '乙型肝炎病毒核心抗体IgM(ELISA)': '乙型肝炎病毒核心抗体IgM', '乙型肝炎病毒核心抗体IgM(CLIA)': '乙型肝炎病毒核心抗体IgM',
    
    # 肿瘤标志物与生化
    '总前列腺特异性抗原(tPSA)': '总前列腺特异性抗原', '总前列腺特异性抗原（tPSA）': '总前列腺特异性抗原',
    '游离前列腺特异性抗原(fPSA)': '游离前列腺特异性抗原', '游离前列腺特异性抗原（fPSA）': '游离前列腺特异性抗原',
    'FPSA/TPSA': 'fPSA/tPSA', 
    '胃蛋白酶原Ⅰ(PGⅠ)': '胃蛋白酶原Ⅰ', '胃蛋白酶原Ⅱ(PGⅡ)': '胃蛋白酶原Ⅱ',
    '细胞角蛋白19(CYFRA21-1)': '非小细胞肺癌相关抗原21-1', '可溶性细胞角蛋白19片段(CYFRA21-1)': '非小细胞肺癌相关抗原21-1', '可溶性细胞角蛋白19片段': '非小细胞肺癌相关抗原21-1',
    
    # 基础生化与脂质
    '低密度脂蛋白': '低密度脂蛋白胆固醇', '高密度脂蛋白': '高密度脂蛋白胆固醇',
    '甘油三脂': '甘油三酯', 
    '胆碱脂酶': '胆碱酯酶', '拟胆碱酯酶': '胆碱酯酶',
    '同型半胱氨酸[HCY]': '同型半胱氨酸',
    '脑脊液蛋白': '脑脊液总蛋白',
    '脂肪酶测定': '脂肪酶',
    '白/球蛋白': '白球比例',
    '血清淀粉样蛋白': '血清淀粉样蛋白A',
    
    # 电解质系列
    '钾测定': '钾', '钾离子': '钾',
    '钠测定': '钠', '钠离子': '钠',
    '氯测定': '氯', '氯离子': '氯',
    
    # 血常规与血型系列
    '白细胞数': '白细胞', '白细胞计数': '白细胞',
    '红细胞数': '红细胞', '红细胞计数': '红细胞',
    '血小板数': '血小板', '血小板计数': '血小板',
    '中性粒细胞数': '中性粒细胞绝对值',
    '单核细胞数': '单核细胞绝对值',
    '淋巴细胞数': '淋巴细胞绝对值',
    '嗜酸性粒细胞数': '嗜酸性粒细胞绝对值',
    '嗜碱性粒细胞数': '嗜碱性粒细胞绝对值',
    '单核细胞比率': '单核细胞百分比', '中性粒细胞比率': '中性粒细胞百分比', '淋巴细胞比率': '淋巴细胞百分比',
    '平均血红蛋白浓度': '平均红细胞血红蛋白浓度',
    '平均血红蛋白含量': '平均红细胞血红蛋白量',
    '异型淋巴细胞（镜检）': '异型淋巴细胞',
    '血小板分布幅': '血小板分布宽度',
    '大血小板': '大型血小板比率', # [V6修复] 灰区提拔
    'Rh分型': 'Rh血型',
    
    # 尿常规、体液
    '红细胞(镜检)': '镜检红细胞', '红细胞（镜检）': '镜检红细胞',
    '白细胞(镜检)': '镜检白细胞', '白细胞（镜检）': '镜检白细胞',
    '尿比重': '比重', '尿比密': '比重', 
    '尿PH值': '尿酸碱度', 'pH值': '尿酸碱度',
    '尿隐血': '隐血',
    '尿葡萄糖': '尿糖',
    '粘液丝': '粘液',
    '亚硝酸盐': '尿亚硝酸盐', # [V6修复] 灰区提拔
    
    # 骨代谢及其他
    '总Ⅰ型(前)胶原氨基端延长肽': '总Ⅰ型胶原氨基端延长肽',
    '血沉（毛细管法）': '血沉', '血沉（光度计法）': '血沉',
    '国际标准化比率': '国际标准化比值', 'UN:CREA': 'BUN:CREA', '尿素氮': '尿素', '尿素计算': '尿素', '肌酐计算': '肌酐',
    
    # 错别字与异体字修复
    'a-L-岩藻糖甘酶': 'a-L-岩藻糖苷酶', '病理性管型': '病理管型', '腊样管型': '蜡样管型',
    '类酵母菌': '酵母菌', '酵母样菌': '酵母菌',
    '胃泌素释放前肽前体': '胃泌素释放肽前体', 
    '谷氨酰转酞酶': '谷氨酰转肽酶', '谷氨酰基转肽酶': '谷氨酰转肽酶',
    '亮氨酸氨基转肽酶': '亮氨酸氨基肽酶', '腺苷酸脱氨酶': '腺苷脱氨酶', 
    '超敏C-反应蛋白': '超敏C反应蛋白', '糖类抗原153': '糖类抗原15-3', '糖类抗原199': '糖类抗原19-9',
    '意外抗体筛查试验': '抗体筛查试验',
    'B-羟丁酸': 'β-羟丁酸'
}

SUSPECT_MERGE_DICT = {
    # [V6修复] 绿区降级，回归黄区让医生定夺
    '葡萄糖': '【存疑待确认】血糖(含空腹与随机)', '空腹血糖': '【存疑待确认】血糖(含空腹与随机)',
    '血淀粉酶': '【存疑待确认】淀粉酶', '淀粉酶': '【存疑待确认】淀粉酶', 'α-淀粉酶': '【存疑待确认】淀粉酶',
    
    # 原有黄区保留
    '肾小球滤过率(MDRD)': '【存疑待确认】肾小球滤过率', '肾小球滤过率(eGFR-EPI)': '【存疑待确认】肾小球滤过率', '肾小球滤过率（女）': '【存疑待确认】肾小球滤过率', '肾小球滤过率(男)': '【存疑待确认】肾小球滤过率',
    '隐血试验（胶体金法）': '【存疑待确认】隐血试验(大便)', '隐血试验': '【存疑待确认】隐血试验(大便)', '隐血试验（匹拉米洞法）': '【存疑待确认】隐血试验(大便)', '隐血试验（化学法）': '【存疑待确认】隐血试验(大便)',
    '高敏肌钙蛋白I': '【存疑待确认】肌钙蛋白I(高敏法可否合并)', '肌钙蛋白I': '【存疑待确认】肌钙蛋白I(高敏法可否合并)',
    '肌酸激酶同工酶（MB）': '【存疑待确认】肌酸激酶同工酶', '肌酸激酶同工酶（MB质量法）': '【存疑待确认】肌酸激酶同工酶', '肌酸激酶同工酶（质量法)': '【存疑待确认】肌酸激酶同工酶', '肌酸激酶同工酶': '【存疑待确认】肌酸激酶同工酶',
    'ROMA值(绝经前)': '【存疑待确认】ROMA值', 'ROMA值(绝经后)': '【存疑待确认】ROMA值'
}

# ================= 2. 数据处理与映射 =================
df = pd.read_csv(input_file)

def extract_analyte(full_name):
    parts = str(full_name).split(' - ')
    if len(parts) >= 2: return ' - '.join(parts[1:]).strip()
    return str(full_name).strip()

df['原始明细'] = df['完整检验项 (项目 - 明细)'].apply(extract_analyte)
df['原始明细'] = df['原始明细'].str.strip(' *-_')
raw_counts = df.groupby('原始明细')['出现总频次'].sum().reset_index()

def get_target_and_cat(name):
    if name in SAFE_MERGE_DICT: return SAFE_MERGE_DICT[name], '1_Safe'
    if name in SUSPECT_MERGE_DICT: return SUSPECT_MERGE_DICT[name], '2_Suspect'
    return name, '3_DoNotMerge'

raw_counts[['初步标准名称', '预设分类']] = raw_counts['原始明细'].apply(lambda x: pd.Series(get_target_and_cat(x)))

result_data = []
for std_name, group in raw_counts.groupby('初步标准名称'):
    if len(group) == 1:
        final_name = group['原始明细'].iloc[0]  
        final_cat = '3_DoNotMerge'           
    else:
        final_name = std_name
        if '2_Suspect' in group['预设分类'].values:
            final_cat = '2_Suspect'
        else:
            final_cat = '1_Safe'

    total_freq = group['出现总频次'].sum()
    details = "  |  ".join([f"{r['原始明细']} (频次:{r['出现总频次']})" for _, r in group.sort_values(by='出现总频次', ascending=False).iterrows()])
    
    result_data.append({
        'Category': final_cat, '推荐标准名称': final_name, '总合并频次': total_freq, 
        '合并项数量': len(group), '包含的原始明细变体': details
    })

final_df = pd.DataFrame(result_data)

# ================= 3. 写入 Excel 并上色 + 渲染框线 =================
wb = Workbook()
ws = wb.active
ws.title = "检验指标靶向审查V6_Final"

headers = ['模块分类', '推荐标准名称', '总合并频次', '合并项数量', '包含的原始明细变体']
ws.append(headers)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

color_safe = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     
color_suspect = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  
color_danger = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   

for cat in ['1_Safe', '2_Suspect', '3_DoNotMerge']:
    sub_df = final_df[final_df['Category'] == cat].sort_values(by=['合并项数量', '总合并频次'], ascending=[False, False])
    if len(sub_df) == 0: continue
    
    if cat == '1_Safe':
        cat_name = "【一、靶向安全合并 (已完美对齐模板病例)】"
        fill_color = color_safe
    elif cat == '2_Suspect':
        cat_name = "【二、存疑待定夺 (存在医学变体，需医生评估是否合并)】"
        fill_color = color_suspect
    else:
        cat_name = "【三、独立指标 (无重复变体或绝对不可混淆项，保持现状)】"
        fill_color = color_danger
        
    ws.append(['', '', '', '', '']) 
    
    for _, row in sub_df.iterrows():
        ws.append([cat_name, row['推荐标准名称'], row['总合并频次'], row['合并项数量'], row['包含的原始明细变体']])
        for cell in ws[ws.max_row]:
            cell.fill = fill_color
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border 

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 100

wb.save(output_file)
print(f"🎉 V6 最终收官版 Excel 生成完毕！已保存至: {output_file}")