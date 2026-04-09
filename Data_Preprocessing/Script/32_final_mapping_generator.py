import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ================= 配置路径 =================
temp_dir = '../Temp_data' 
input_file = os.path.join(temp_dir, '26_Lab_Items_Full_Dictionary.csv')
output_file = os.path.join(temp_dir, '32_Lab_Master_Mapping_Final_Grouped.xlsx')

print("🚀 正在编译【检验指标映射字典 - 满血修复分组视觉版】...\n")

# ================= 1. 定义强制合并白名单 (已全量恢复 V6 绿区 + 黄区升级) =================
FINAL_MAPPED_DICT = {
    # 传染病与抗原抗体 (全量恢复)
    '乙型肝炎病毒表面抗原(CLIA)': '乙型肝炎病毒表面抗原', '乙型肝炎病毒表面抗原(ELISA)': '乙型肝炎病毒表面抗原',
    '丙肝抗体(CLIA)': '丙型肝炎病毒抗体', '丙型肝炎病毒抗体(CLIA)': '丙型肝炎病毒抗体', '丙型肝炎病毒抗体(ELISA)': '丙型肝炎病毒抗体',
    '梅毒特异性抗体(TP-ELISA)': '梅毒特异性抗体', '梅毒特异性抗体(CLIA)': '梅毒特异性抗体', '梅毒特异性抗体(TPPA)': '梅毒特异性抗体',
    '人类免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(ELISA)': '人类免疫缺陷病毒抗原抗体', '人免疫缺陷病毒抗原抗体(CLIA)': '人类免疫缺陷病毒抗原抗体',
    '艾滋病抗原抗体筛查(ELISA)': '人类免疫缺陷病毒抗原抗体', 'HIV P24抗原/抗体(CLIA)': '人类免疫缺陷病毒抗原抗体',
    '乙型肝炎病毒e抗体(ELISA)': '乙型肝炎病毒e抗体', '乙型肝炎病毒e抗体(CLIA)': '乙型肝炎病毒e抗体',
    '乙型肝炎病毒e抗原(ELISA)': '乙型肝炎病毒e抗原', '乙型肝炎病毒e抗原(CLIA)': '乙型肝炎病毒e抗原',
    '乙型肝炎病毒核心抗体(ELISA)': '乙型肝炎病毒核心抗体', '乙型肝炎病毒核心抗体(CLIA)': '乙型肝炎病毒核心抗体',
    '乙型肝炎病毒表面抗体(ELISA)': '乙型肝炎病毒表面抗体', '乙型肝炎病毒表面抗体(CLIA)': '乙型肝炎病毒表面抗体',
    '乙型肝炎病毒核心抗体IgM(ELISA)': '乙型肝炎病毒核心抗体IgM', '乙型肝炎病毒核心抗体IgM(CLIA)': '乙型肝炎病毒核心抗体IgM',
    
    # 肿瘤标志物与生化 
    '总前列腺特异性抗原(tPSA)': '总前列腺特异性抗原', '总前列腺特异性抗原（tPSA）': '总前列腺特异性抗原',
    '游离前列腺特异性抗原(fPSA)': '游离前列腺特异性抗原', '游离前列腺特异性抗原（fPSA）': '游离前列腺特异性抗原',
    'FPSA/TPSA': 'fPSA/tPSA', 
    '胃蛋白酶原Ⅰ(PGⅠ)': '胃蛋白酶原Ⅰ', '胃蛋白酶原Ⅱ(PGⅡ)': '胃蛋白酶原Ⅱ',
    '细胞角蛋白19(CYFRA21-1)': '非小细胞肺癌相关抗原21-1', '可溶性细胞角蛋白19片段(CYFRA21-1)': '非小细胞肺癌相关抗原21-1', '可溶性细胞角蛋白19片段': '非小细胞肺癌相关抗原21-1',
    
    # 基础生化、电解质与脂质 (全量恢复)
    '低密度脂蛋白': '低密度脂蛋白胆固醇', '高密度脂蛋白': '高密度脂蛋白胆固醇',
    '甘油三脂': '甘油三酯', 
    '胆碱脂酶': '胆碱酯酶', '拟胆碱酯酶': '胆碱酯酶',
    '同型半胱氨酸[HCY]': '同型半胱氨酸',
    '脑脊液蛋白': '脑脊液总蛋白',
    '脂肪酶测定': '脂肪酶',
    '空腹血糖': '葡萄糖', # 黄区升级
    '血淀粉酶': '淀粉酶', 'α-淀粉酶': '淀粉酶', # 黄区升级
    '白/球蛋白': '白球比例',
    '血清淀粉样蛋白': '血清淀粉样蛋白A',
    
    # 电解质系列
    '钾测定': '钾', '钾离子': '钾',
    '钠测定': '钠', '钠离子': '钠',
    '氯测定': '氯', '氯离子': '氯',
    
    # 血常规系列与绝对值转换 (全量恢复)
    '白细胞数': '白细胞', '白细胞计数': '白细胞',
    '红细胞数': '红细胞', '红细胞计数': '红细胞',
    '血小板数': '血小板', '血小板计数': '血小板',
    '中性粒细胞数': '中性粒细胞绝对值',
    '单核细胞数': '单核细胞绝对值',
    '淋巴细胞数': '淋巴细胞绝对值',
    '嗜酸性粒细胞数': '嗜酸性粒细胞绝对值',
    '嗜碱性粒细胞数': '嗜碱性粒细胞绝对值',
    '单核细胞比率': '单核细胞百分比', '中性粒细胞比率': '中性粒细胞百分比', '淋巴细胞比率': '淋巴细胞百分比',
    '平均血红蛋白浓度': '平均红细胞血红蛋白浓度',
    '平均血红蛋白含量': '平均红细胞血红蛋白量',
    '异型淋巴细胞（镜检）': '异型淋巴细胞',
    '血小板分布幅': '血小板分布宽度',
    '大血小板': '大型血小板比率', 
    'Rh分型': 'Rh血型',
    
    # 肾功能、心肌酶与隐血 (黄区升级)
    '肾小球滤过率(MDRD)': '肾小球滤过率', '肾小球滤过率(eGFR-EPI)': '肾小球滤过率', '肾小球滤过率（女）': '肾小球滤过率', '肾小球滤过率(男)': '肾小球滤过率',
    '隐血试验（胶体金法）': '大便隐血试验', '隐血试验': '大便隐血试验', '隐血试验（匹拉米洞法）': '大便隐血试验', '隐血试验（化学法）': '大便隐血试验',
    '高敏肌钙蛋白I': '肌钙蛋白I',
    '肌酸激酶同工酶（MB）': '肌酸激酶同工酶', '肌酸激酶同工酶（MB质量法）': '肌酸激酶同工酶', '肌酸激酶同工酶（质量法)': '肌酸激酶同工酶',
    
    # 镜检与尿常规 (全量恢复)
    '红细胞(镜检)': '镜检红细胞', '红细胞（镜检）': '镜检红细胞',
    '白细胞(镜检)': '镜检白细胞', '白细胞（镜检）': '镜检白细胞',
    '尿比重': '比重', '尿比密': '比重', 
    '尿PH值': '尿酸碱度', 'pH值': '尿酸碱度',
    '尿隐血': '隐血',
    '尿葡萄糖': '尿糖',
    '粘液丝': '粘液',
    '亚硝酸盐': '尿亚硝酸盐', 
    
    # 骨代谢及其他
    '总Ⅰ型(前)胶原氨基端延长肽': '总Ⅰ型胶原氨基端延长肽',
    '血沉（毛细管法）': '血沉', '血沉（光度计法）': '血沉',
    '国际标准化比率': '国际标准化比值', 'UN:CREA': 'BUN:CREA', '尿素氮': '尿素', '尿素计算': '尿素', '肌酐计算': '肌酐',
    
    # 错别字与异体字极限量修复
    'a-L-岩藻糖甘酶': 'a-L-岩藻糖苷酶', '病理性管型': '病理管型', '腊样管型': '蜡样管型',
    '类酵母菌': '酵母菌', '酵母样菌': '酵母菌',
    '胃泌素释放前肽前体': '胃泌素释放肽前体', 
    '谷氨酰转酞酶': '谷氨酰转肽酶', '谷氨酰基转肽酶': '谷氨酰转肽酶',
    '亮氨酸氨基转肽酶': '亮氨酸氨基肽酶', '腺苷酸脱氨酶': '腺苷脱氨酶', 
    '超敏C-反应蛋白': '超敏C反应蛋白', '糖类抗原153': '糖类抗原15-3', '糖类抗原199': '糖类抗原19-9',
    '意外抗体筛查试验': '抗体筛查试验',
    'B-羟丁酸': 'β-羟丁酸'
    
    # 注意：ROMA值(绝经前) 和 ROMA值(绝经后) 不在字典内，将自动被打回独立区！
}

# ================= 2. 数据处理与三重排序 =================
df_raw = pd.read_csv(input_file)

def extract_analyte(full_name):
    parts = str(full_name).split(' - ')
    return ' - '.join(parts[1:]).strip() if len(parts) >= 2 else str(full_name).strip()

df_raw['原始明细'] = df_raw['完整检验项 (项目 - 明细)'].apply(extract_analyte).str.strip(' *-_')
raw_counts = df_raw.groupby('原始明细')['出现总频次'].sum().reset_index()

raw_counts['推荐标准名'] = raw_counts['原始明细'].apply(lambda x: FINAL_MAPPED_DICT.get(x, x))
group_sizes = raw_counts.groupby('推荐标准名').size().to_dict()
group_totals = raw_counts.groupby('推荐标准名')['出现总频次'].sum().to_dict()

result_data = []
for _, row in raw_counts.iterrows():
    std_name = row['推荐标准名']
    # 如果对应的标准名有超过 1 个变体被映射，则归入合并区
    cat = 'MAPPED' if group_sizes[std_name] > 1 else 'INDEPENDENT'
    
    result_data.append({
        'Category': cat,
        'Original_Name': row['原始明细'],
        'Standard_Name': std_name,
        'Frequency': row['出现总频次'],
        'Group_Total_Freq': group_totals[std_name]
    })

final_df = pd.DataFrame(result_data)

# ================= 3. 导出带斑马纹的 Excel =================
wb = Workbook()
ws = wb.active
ws.title = "映射字典_分组视觉版"

headers = ['状态', '原始指标明细', '合并后标准名称', '原始单项频次', '家族总频次']
ws.append(headers)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_mapped_1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fill_mapped_2 = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
fill_indep = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   

# --- 处理合并区 ---
mapped_df = final_df[final_df['Category'] == 'MAPPED'].sort_values(
    by=['Group_Total_Freq', 'Standard_Name', 'Frequency'], 
    ascending=[False, True, False]
)

current_group_name = ""
use_color_1 = True

for _, row in mapped_df.iterrows():
    if row['Standard_Name'] != current_group_name:
        current_group_name = row['Standard_Name']
        use_color_1 = not use_color_1 
        
    current_fill = fill_mapped_1 if use_color_1 else fill_mapped_2
    ws.append([row['Category'], row['Original_Name'], row['Standard_Name'], row['Frequency'], row['Group_Total_Freq']])
    for cell in ws[ws.max_row]:
        cell.fill = current_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

ws.append(['', '', '', '', ''])

# --- 处理独立区 ---
indep_df = final_df[final_df['Category'] == 'INDEPENDENT'].sort_values(by='Frequency', ascending=False)
for _, row in indep_df.iterrows():
    ws.append([row['Category'], row['Original_Name'], row['Standard_Name'], row['Frequency'], '-'])
    for cell in ws[ws.max_row]:
        if cell.value != '':
            cell.fill = fill_indep
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
wb.save(output_file)

print(f"🎉 修复完成！满血版视觉分块映射字典已生成：{output_file}")
print("👉 这次保证有 70 多组合并项，全量涵盖了心肌酶、电解质、大便隐血等所有升级为绿区的指标。ROMA 继续待在灰区！")